from fastapi import FastAPI
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.drawing.image import Image

import requests
import uuid
import os

app = FastAPI()

# {
#   "asin": "B0G7XTH8HR",
#   "brand": "UNA GELLA",
#   "brandUrl": "https://www.amazon.com/s?k=UNA GELLA",
#   "imageUrl": "https://images-na.ssl-images-amazon.com/images/I/81IP0thbSoL._AC_US200_.jpg",
#   "title": "UNA GELLA Oval Nail Tips Exra Short 504PCS Pre-shape Gel Oval Almond Nail Tips for Women DIY Salon Round Almond Press On Nails Full Cover Clear Fake Acrylic Tips 12 Sizes",
#   "parent": "B0GWXN9WTC",
#   "nodeLabelPath": "Beauty & Personal Care:Foot, Hand & Nail Care:Nail Art & Polish:False Nails & Accessories:Nail Tips",
#   "nodeIdPath": "3760911:17242866011:11059311:16261986011:13861684011",
#   "nodeId": 13861684011,
#   "bsrId": "beauty",
#   "bsr": 95,
#   "bsrCv": 15,
#   "bsrCr": 13.64,
#   "amzUnit": 600,
#   "amzUnitDate": 1777528565000,
#   "amzSales": 5988,
#   "units": 89533,
#   "unitsGr": null,
#   "revenue": 894434.6,
#   "price": 9.98,
#   "averagePrice": 9.99,
#   "primePrice": -1,
#   "profit": 58.48,
#   "fba": 2.65,
#   "ratings": 16382,
#   "ratingsRate": 0.76,
#   "rating": 4.5,
#   "ratingsCv": 682,
#   "ratingDelta": 94,
#   "availableDate": 1770163200000,
#   "fulfillment": "FBA",
#   "variations": 22,
#   "sellers": 1,
#   "sellerName": "UNA GELLA STORE",
#   "sellerId": "AT5KPP94U9F48",
#   "sellerNation": "CN",
#   "lqs": 96,
#   "weight": "4.59 ounces",
#   "dimension": "8.58 x 5 x 0.67 inches",
#   "pkgDimensions": "8.6 x 5 x 0.7 inches",
#   "pkgDimensionType": "SS",
#   "pkgWeight": "0.29 pounds",
#   "sku": "Pattern Name: Oval XS",
#   "dimensionsType": "SS",
#   "deliveryPrice": -1,
#   "badge": {
#     "bestSeller": "#1 Best Seller  in False Nail Tips",
#     "amazonChoice": "N",
#     "newRelease": "N",
#     "ebc": "Y",
#     "video": "N"
#   },
#   "subcategories": [
#     {
#       "code": "13861684011",
#       "rank": 1,
#       "label": "False Nail Tips"
#     }
#   ],
#   "symbol": "N"
# }

class SubCategory(BaseModel):
    code: str
    rank: int
    label: str

class Badge(BaseModel):
    bestSeller: str
    amazonChoice: str
    newRelease: str
    ebc: str
    video: str

class Item(BaseModel):
    asin: str
    brand: str
    brandUrl: str
    imageUrl: str
    title: str
    parent: str
    nodeLabelPath: str
    nodeIdPath: str
    nodeId: int
    bsrId: str
    bsr: int
    bsrCv: int
    bsrCr: float
    amzUnit: int
    amzUnitDate: str
    amzSales: int
    units: int
    unitsGr: int | None
    revenue: float
    price: float
    averagePrice: float
    primePrice: float
    profit: float
    fba: float
    ratings: int
    ratingsRate: float
    rating: float
    ratingsCv: int
    ratingDelta: int
    availableDate: str
    fulfillment: str
    variations: int
    sellers: int
    sellerName: str
    sellerId: str
    sellerNation: str
    lqs: int
    weight: str
    dimension: str
    pkgDimensions: str
    pkgDimensionType: str
    pkgWeight: str
    sku: str
    dimensionsType: str
    deliveryPrice: float
    badge: Badge
    subcategories: list[SubCategory]
    symbol: str

@app.post("/xlsx")
async def create_xlsx(items: list[Item]):

    wb = Workbook()
    ws = wb.active

    ws.append(["asin", "brand", "brandUrl", "imageUrl", "title", "parent", "nodeLabelPath", "nodeIdPath", "nodeId", "bsrId", "bsr", "bsrCv", "bsrCr", "amzUnit", "amzUnitDate", "amzSales", "units", "unitsGr", "revenue", "price", "averagePrice", "primePrice", "profit", "fba", "ratings", "ratingsRate", "rating", "ratingsCv", "ratingDelta", "availableDate", "fulfillment", "variations", "sellers", "sellerName", "sellerId", "sellerNation", "lqs", "weight", "dimension", "pkgDimensions", "pkgDimensionType", "pkgWeight", "sku", "dimensionsType", "deliveryPrice", "badge", "subcategories", "symbol"])

    for idx, item in enumerate(items, start=2):

        ws.cell(idx, 1, item.asin)
        ws.cell(idx, 2, item.brand)
        ws.cell(idx, 3, item.brandUrl)
        # ws.cell(idx, 4, item.imageUrl)

        img_data = requests.get(item.imageUrl).content

        img_path = f"/tmp/{uuid.uuid4()}.jpg"

        with open(img_path, "wb") as f:
            f.write(img_data)

        img = Image(img_path)

        img.width = 100
        img.height = 100

        ws.add_image(img, f"D{idx}")

        ws.row_dimensions[idx].height = 80

        ws.cell(idx, 5, item.title)
        ws.cell(idx, 6, item.parent)
        ws.cell(idx, 7, item.nodeLabelPath)
        ws.cell(idx, 8, item.nodeIdPath)
        ws.cell(idx, 9, item.nodeId)
        ws.cell(idx, 10, item.bsrId)
        ws.cell(idx, 11, item.bsr)
        ws.cell(idx, 12, item.bsrCv)
        ws.cell(idx, 13, item.bsrCr)
        ws.cell(idx, 14, item.amzUnit)
        ws.cell(idx, 15, item.amzUnitDate)
        ws.cell(idx, 16, item.amzSales)
        ws.cell(idx, 17, item.units)
        ws.cell(idx, 18, item.unitsGr)
        ws.cell(idx, 19, item.revenue)
        ws.cell(idx, 20, item.price)
        ws.cell(idx, 21, item.averagePrice)
        ws.cell(idx, 22, item.primePrice)
        ws.cell(idx, 23, item.profit)
        ws.cell(idx, 24, item.fba)
        ws.cell(idx, 25, item.ratings)
        ws.cell(idx, 26, item.ratingsRate)
        ws.cell(idx, 27, item.rating)
        ws.cell(idx, 28, item.ratingsCv)
        ws.cell(idx, 29, item.ratingDelta)
        ws.cell(idx, 30, item.availableDate)
        ws.cell(idx, 31, item.fulfillment)
        ws.cell(idx, 32, item.variations)
        ws.cell(idx, 33, item.sellers)
        ws.cell(idx, 34, item.sellerName)
        ws.cell(idx, 35, item.sellerId)
        ws.cell(idx, 36, item.sellerNation)
        ws.cell(idx, 37, item.lqs)
        ws.cell(idx, 38, item.weight)
        ws.cell(idx, 39, item.dimension)
        ws.cell(idx, 40, item.pkgDimensions)
        ws.cell(idx, 41, item.pkgDimensionType)
        ws.cell(idx, 42, item.pkgWeight)
        ws.cell(idx, 43, item.sku)
        ws.cell(idx, 44, item.dimensionsType)
        ws.cell(idx, 45, item.deliveryPrice)
        ws.cell(idx, 46, f"{item.badge.bestSeller}, {item.badge.amazonChoice}, {item.badge.newRelease}, {item.badge.ebc}, {item.badge.video}")
        ws.cell(idx, 47, ", ".join([f"{sub.label} ({sub.code})" for sub in item.subcategories]))
        ws.cell(idx, 48, item.symbol)

    file_path = "/tmp/products.xlsx"

    wb.save(file_path)

    return FileResponse(
        file_path,
        filename="products.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )